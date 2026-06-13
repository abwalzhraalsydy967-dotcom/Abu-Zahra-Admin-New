import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "المراجعة الوظيفية"
ws.sheet_view.rightToLeft = True

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
subtitle_font = Font(name='Arial', bold=True, size=11, color='333333')
ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
warn_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
err_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ok_font = Font(name='Arial', size=10, color='006100')
warn_font = Font(name='Arial', size=10, color='9C5700')
err_font = Font(name='Arial', size=10, color='9C0006')
cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
cat_font = Font(name='Arial', bold=True, size=10, color='1F4E79')
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

# Title
ws.merge_cells('A1:L1')
ws['A1'] = 'جدول المراجعة الوظيفية الشامل - مشروع أبو الزهراء v3.4'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('A2:L2')
ws['A2'] = 'تاريخ المراجعة: 2026-06-14 | إجمالي الأوامر: 170+ | المكونات: Server + Web Dashboard + Telegram Bot + Android Client + Admin App'
ws['A2'].font = Font(name='Arial', size=9, color='666666')
ws['A2'].alignment = Alignment(horizontal='center')

# Headers
headers = [
    'التصنيف', 'اسم الأمر', 'الأمر الفعلي (cmd)', 'الملف المسؤول',
    'API / WebSocket', 'يصل للجهاز؟', 'ينفذ فعلياً؟',
    'تعود للسيرفر؟', 'تظهر في لوحة التحكم؟', 'تظهر في تطبيق الإدارة؟',
    'تظهر في البوت؟', 'الحالة'
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Command data
commands = [
    # Data Collection
    ('data', 'الرسائل SMS', 'get_sms', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'سجل المكالمات', 'get_calls', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'جهات الاتصال', 'get_contacts', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'الموقع الجغرافي', 'get_location', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'الإشعارات', 'get_notifications', 'DataCollector.kt / MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'التطبيقات المثبتة', 'get_apps', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'معلومات الجهاز', 'get_info', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'حالة البطارية', 'get_battery', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'المعرض', 'get_gallery', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'الحافظة', 'get_clipboard', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'جميع البيانات', 'get_all', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'معلومات الواي فاي', 'get_wifi_info', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'معلومات الشبكة', 'get_network_info', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'معلومات الشريحة', 'get_sim_info', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'معلومات التخزين', 'get_storage_info', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'التطبيقات النشطة', 'get_running_apps', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'التقويم', 'get_calendar', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'سجل المتصفح', 'get_browser_history', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '⚠️ محدود', '✅', '✅', '✅', '✅', '⚠️'),
    ('data', 'التطبيقات المثبتة (تفصيلي)', 'get_installed_apps', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('data', 'استخدام التطبيقات', 'get_app_usage', 'DataCollector.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),

    # Social Media (NOW FIXED)
    ('social', 'واتساب', 'get_whatsapp', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'تليجرام', 'get_telegram', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'انستجرام', 'get_instagram', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'ماسنجر', 'get_messenger', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'سناب شات', 'get_snapchat', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'تيك توك', 'get_tiktok', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'تويتر / X', 'get_twitter', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'فايبر', 'get_viber', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'سيجنال', 'get_signal', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'فيسبوك', 'get_facebook', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'يوتيوب', 'get_tiktok', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'حالات واتساب', 'get_whatsapp', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'قصص واتساب', 'get_whatsapp', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'قنوات تليجرام', 'get_telegram', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('social', 'قصص انستجرام', 'get_instagram', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),

    # Remote Control
    ('control', 'فحص الاتصال', 'ping', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'اهتزاز', 'vibrate', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'رنين', 'ring', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'لقطة شاشة', 'screenshot', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase + /api/upload', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'كاميرا أمامية', 'front_camera', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase + /api/upload', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'كاميرا خلفية', 'back_camera', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase + /api/upload', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تسجيل صوتي', 'record_audio', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase + /api/upload', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تسجيل فيديو', 'record_screen', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase + /api/upload', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'قفل الهاتف', 'lock_phone', 'ControlExecutor.kt / SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'فتح الهاتف', 'unlock_phone', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'إعادة تشغيل', 'reboot', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'إيقاف التشغيل', 'shutdown', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '⚠️ يحتاج Device Admin', '✅', '✅', '✅', '✅', '⚠️'),
    ('control', 'تعيين الصوت', 'set_volume', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تعيين السطوع', 'set_brightness', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تشغيل الواي فاي', 'enable_wifi', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'إيقاف الواي فاي', 'disable_wifi', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تشغيل البلوتوث', 'enable_bluetooth', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'إيقاف البلوتوث', 'disable_bluetooth', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تشغيل بيانات الجوال', 'enable_mobile_data', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '⚠️ يحتاج صلاحية', '✅', '✅', '✅', '✅', '⚠️'),
    ('control', 'إيقاف بيانات الجوال', 'disable_mobile_data', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '⚠️ يحتاج صلاحية', '✅', '✅', '✅', '✅', '✅', '⚠️'),
    ('control', 'تشغيل نقطة الاتصال', 'enable_hotspot', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'وضع الطيار تشغيل', 'airplane_on', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'وضع الطيار إيقاف', 'airplane_off', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تشغيل الكشاف', 'torch_on', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'إطفاء الكشاف', 'torch_off', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تشغيل صوت', 'play_sound', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'نطق نص', 'speak_text', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'إظهار إشعار', 'show_notification', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'فتح رابط', 'open_url', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'إرسال SMS', 'send_sms', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'إجراء مكالمة', 'make_call', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'حظر رقم', 'block_number', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'إلغاء حظر رقم', 'unblock_number', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تعيين النغمة', 'set_ringtone', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'تعيين الخلفية', 'set_wallpaper', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('control', 'الدوران التلقائي', 'set_auto_rotate', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),

    # App Management
    ('apps', 'فتح تطبيق', 'open_app', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'إغلاق تطبيق', 'close_app', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'تثبيت تطبيق', 'install_app', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'حذف تطبيق', 'uninstall_app', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'حظر تطبيق', 'block_app', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'إلغاء حظر تطبيق', 'unblock_app', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'مسح بيانات تطبيق', 'clear_app_data', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'إيقاف قسري', 'force_stop_app', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'معلومات تطبيق', 'app_info', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'صلاحيات التطبيق', 'app_permissions', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'تفعيل تطبيق', 'enable_app', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'تعطيل تطبيق', 'disable_app', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'وقت الشاشة', 'get_app_usage', 'AppExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'قائمة المحظورة', 'list_blocked', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('apps', 'ضغط ملفات', 'zip_files', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),

    # File Management
    ('files', 'عرض الملفات', 'list_files', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'جلب ملف', 'get_file', 'FileExecutor.kt + /api/upload', 'POST /api/web/send_command + /api/upload', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'حذف ملف', 'delete_file', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'إعادة تسمية', 'rename_file', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'نسخ ملف', 'copy_file', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'نقل ملف', 'move_file', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'إنشاء مجلد', 'create_folder', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'بحث في الملفات', 'search_files', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'الملفات الأخيرة', 'recent_files', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'معلومات ملف', 'file_info', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('files', 'حجم المجلد', 'get_folder_size', 'FileExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),

    # Security
    ('security', 'مسح البيانات', 'wipe_data', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('security', 'إعادة ضبط المصنع', 'factory_reset', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('security', 'إظهار التطبيق', 'show_app', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('security', 'إخفاء التطبيق', 'hide_app', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('security', 'تغيير رمز القفل', 'change_passcode', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('security', 'تشغيل البصمة', 'enable_biometric', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('security', 'إيقاف البصمة', 'disable_biometric', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('security', 'الحماية من الحذف', 'anti_uninstall_on', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('security', 'حالة مسؤول الجهاز', 'device_admin_status', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('security', 'فحص الروت', 'get_info', 'SecurityExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),

    # Monitoring
    ('monitor', 'بدء تسجيل المفاتيح', 'keylogger_start', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'إيقاف تسجيل المفاتيح', 'keylogger_stop', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'جلب بيانات المفاتيح', 'get_keylogger', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'بدء تسجيل الشاشة', 'screen_record_start', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'إيقاف تسجيل الشاشة', 'stop_screen', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'تتبع مباشر', 'location_live', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'إيقاف التتبع', 'location_stop', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'مراقبة الحافظة', 'clipboard_monitor_start', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'مراقبة الواي فاي', 'wifi_monitor_start', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'مراقبة التطبيقات', 'app_monitor_start', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'مراقبة الرسائل', 'sms_monitor', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('monitor', 'مراقبة المكالمات', 'call_monitor', 'MonitorExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),

    # Streaming
    ('streaming', 'بث الشاشة', 'start_screen_stream', 'StreamExecutor.kt', 'POST /api/web/send_command + /ws/stream', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('streaming', 'إيقاف بث الشاشة', 'stop_screen_stream', 'StreamExecutor.kt', 'WebSocket', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('streaming', 'بث الكاميرا', 'start_camera_stream', 'StreamExecutor.kt', 'POST /api/web/send_command + /ws/stream', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('streaming', 'بث الصوت', 'start_audio_stream', 'StreamExecutor.kt', 'POST /api/web/send_command + /ws/stream', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('streaming', 'تبديل الكاميرا', 'switch_camera', 'StreamExecutor.kt', 'WebSocket', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('streaming', 'حالة البث', 'get_stream_status', 'StreamExecutor.kt', 'GET /api/stream/status', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('streaming', 'إيقاف كل البث', 'stop_all_streams', 'StreamExecutor.kt', 'WebSocket', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),

    # System Settings (NOW FIXED)
    ('syssettings', 'تعيين اللغة', 'set_language', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('syssettings', 'تعيين المنطقة الزمنية', 'set_timezone', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('syssettings', 'تعيين منبه', 'set_alarm', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('syssettings', 'وضع المطور', 'enable_dev_mode', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('syssettings', 'تصحيح USB', 'enable_usb_debug', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('syssettings', 'تغيير DNS', 'dns_change', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('syssettings', 'إعدادات APN', 'apn_settings', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('syssettings', 'NFC تشغيل', 'nfc_on', 'ControlExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
    ('syssettings', 'التحديث التلقائي', 'auto_update_on', 'CommandExecutor.kt', 'POST /api/web/send_command + Firebase', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅'),
]

# Write data
row = 5
prev_cat = None
for cmd in commands:
    cat = cmd[0]
    # Category separator row
    if cat != prev_cat:
        cat_names = {
            'data': '📊 جمع البيانات',
            'social': '🌐 التواصل الاجتماعي',
            'control': '🎮 التحكم عن بعد',
            'apps': '📦 إدارة التطبيقات',
            'files': '📂 إدارة الملفات',
            'security': '🔒 الأمان والإدارة',
            'monitor': '🔍 المراقبة',
            'streaming': '📡 البث المباشر',
            'syssettings': '⚙️ إعدادات النظام',
        }
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        cell = ws.cell(row=row, column=1, value=cat_names.get(cat, cat))
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(horizontal='right', vertical='center')
        for c in range(1, 13):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = cat_fill
        row += 1
        prev_cat = cat

    for col, val in enumerate(cmd, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = center_align if col > 4 else wrap_align
        cell.border = thin_border
        cell.font = Font(name='Arial', size=9)

        # Status column coloring
        if col == 12:
            if val == '✅':
                cell.fill = ok_fill
                cell.font = ok_font
            elif '⚠️' in str(val):
                cell.fill = warn_fill
                cell.font = warn_font
            elif val == '❌':
                cell.fill = err_fill
                cell.font = err_font

        # Check/Fail coloring for columns 6-11
        if col >= 6 and col <= 11:
            if val == '✅':
                cell.fill = ok_fill
                cell.font = Font(name='Arial', size=9, color='006100')
            elif '⚠️' in str(val):
                cell.fill = warn_fill
                cell.font = Font(name='Arial', size=9, color='9C5700')
            elif val == '❌':
                cell.fill = err_fill
                cell.font = Font(name='Arial', size=9, color='9C0006')

    row += 1

# Column widths
widths = [18, 22, 25, 28, 35, 14, 14, 14, 18, 18, 16, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Summary sheet
ws2 = wb.create_sheet("ملخص الإصلاحات")
ws2.sheet_view.rightToLeft = True

ws2.merge_cells('A1:E1')
ws2['A1'] = 'ملخص الإصلاحات المطبقة'
ws2['A1'].font = title_font

fixes = [
    ('تطبيق الإدارة', 'AdminApp.kt غير موجود', 'تم إنشاء الملف', 'حرج'),
    ('تطبيق الإدارة', 'LogsActivity.kt يحتوي كود مكرر', 'تم إعادة كتابة الملف بالكامل', 'حرج'),
    ('تطبيق الإدارة', 'SettingsActivity.kt غير موجود', 'تم إنشاء الملف بالكامل', 'حرج'),
    ('تطبيق الإدارة', 'FileProvider غير معرف في Manifest', 'تم إضافة FileProvider + file_paths.xml', 'حرج'),
    ('تطبيق الإدارة', 'نقاط النهاية API خاطئة', 'تم تصحيحها لتتطابق مع السيرفر', 'حرج'),
    ('تطبيق الإدارة', 'أسماء الأوامر لا تتطابق مع السيرفر', 'تم تحديث جميع الأوامر (79 أمر)', 'حرج'),
    ('تطبيق الإدارة', 'فصول بيانات مفقودة', 'تم إنشاء SendCommandRequest/Response/RemoteFile/CommandDefinitions', 'حرج'),
    ('تطبيق العميل', 'أوامر التواصل الاجتماعي (9) stub', 'تم استبدالها بجمع بيانات حقيقي من التخزين', 'عالي'),
    ('تطبيق العميل', 'enable_app / disable_app غير مُنفذ', 'تم تنفيذها عبر PackageManager', 'عالي'),
    ('تطبيق العميل', 'app_permissions غير مُنفذ', 'تم تنفيذها عبر GET_PERMISSIONS', 'عالي'),
    ('تطبيق العميل', 'zip_files غير مُنفذ', 'تم تنفيذها بـ java.util.zip', 'عالي'),
    ('تطبيق العميل', 'unlock_phone غير مُنفذ', 'تم تنفيذها عبر AccessibilityService', 'عالي'),
    ('تطبيق العميل', 'set_wallpaper غير مُنفذ', 'تم تنفيذها عبر WallpaperManager', 'عالي'),
    ('تطبيق العميل', 'block/unblock_number غير مُنفذ', 'تم تنفيذها عبر BlockedNumberContract', 'عالي'),
    ('تطبيق العميل', 'list_blocked غير مُنفذ', 'تم تنفيذها عبر ContentResolver', 'عالي'),
    ('تطبيق العميل', 'enable_dev_mode وغيرها رسائل ثابتة', 'تم تنفيذها لفتح الإعدادات فعلياً', 'متوسط'),
]

fix_headers = ['المكون', 'المشكلة', 'الإصلاح', 'الأولوية']
for col, h in enumerate(fix_headers, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for i, fix in enumerate(fixes, 4):
    for j, val in enumerate(fix, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        cell.alignment = wrap_align if j < 4 else center_align
        cell.border = thin_border
        cell.font = Font(name='Arial', size=10)
        if j == 4:
            if val == 'حرج':
                cell.fill = err_fill
                cell.font = Font(name='Arial', size=10, bold=True, color='9C0006')
            elif val == 'عالي':
                cell.fill = warn_fill
            else:
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 12

# Save
path = '/home/z/my-project/download/functional_review.xlsx'
wb.save(path)
print(f'Saved to {path}')
